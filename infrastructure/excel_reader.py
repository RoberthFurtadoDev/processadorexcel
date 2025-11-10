"""Adaptador de leitura Excel."""
from typing import List, Optional
from openpyxl import load_workbook
from domain.entities import Imovel
from domain.services import FormatadorNumerico, FormatadorRotulos


class ExcelReader:
    """Lê dados de imóveis de planilhas Excel."""
    
    IDX = {
        "EMPREEND": 0, "TIPO": 1, "AREA_TOTAL": 3, "AREA_PRIV": 6,
        "AREA_TERR": 9, "QTD_QUARTO": 11, "LBL_QUARTO": 12, "VAGA": 13,
        "QTD_VARANDA": 14, "LBL_VARANDA": 15, "QTD_ASERV": 16, "LBL_ASERV": 17,
        "QTD_BANHEIRO": 18, "LBL_BANHEIRO": 19, "QTD_SALA": 20, "LBL_SALA": 21,
        "QTD_COZINHA": 22, "LBL_COZINHA": 23, "DESCRICAO": 24,
        "IPTU_VAL": 26, "MATRICULA": 28, "OFICIO": 30,
    }
    
    @staticmethod
    def _cell_text(c):
        v = c.value if hasattr(c, "value") else c
        return "" if v is None else str(v).strip()
    
    @staticmethod
    def _pick_worksheet(wb, preferred_name: Optional[str]):
        names = wb.sheetnames
        if preferred_name and preferred_name in names:
            return wb[preferred_name]
        if preferred_name:
            pref = preferred_name.strip().lower()
            for n in names:
                if n.strip().lower() == pref or pref in n.strip().lower():
                    return wb[n]
        return wb.active
    
    def ler_imoveis(self, input_path: str, sheet_name: str) -> List[Imovel]:
        wb = load_workbook(input_path, data_only=True, read_only=True)
        ws = self._pick_worksheet(wb, sheet_name)
        
        imoveis = []
        first = True
        
        for row in ws.iter_rows(values_only=False):
            if first:
                first = False
                continue
            
            if not row or all(self._cell_text(c) == "" for c in row):
                continue
            
            imovel = self._parse_row(row)
            if imovel:
                imoveis.append(imovel)
        
        wb.close()
        return imoveis
    
    def _parse_row(self, row) -> Optional[Imovel]:
        fmt = FormatadorNumerico()
        lbl = FormatadorRotulos()
        
        return Imovel(
            empreendimento=self._cell_text(row[self.IDX["EMPREEND"]]),
            tipo=self._cell_text(row[self.IDX["TIPO"]]),
            area_total=fmt.format_decimal_pt(self._cell_text(row[self.IDX["AREA_TOTAL"]]), 2),
            area_privativa=fmt.format_decimal_pt(self._cell_text(row[self.IDX["AREA_PRIV"]]), 2),
            area_terreno=fmt.format_decimal_pt(self._cell_text(row[self.IDX["AREA_TERR"]]), 2),
            quartos=lbl.pair_qty_label(
                self._cell_text(row[self.IDX["QTD_QUARTO"]]),
                self._cell_text(row[self.IDX["LBL_QUARTO"]]),
                "quartos"
            ),
            vaga=self._cell_text(row[self.IDX["VAGA"]]),
            varanda=lbl.pair_qty_label(
                self._cell_text(row[self.IDX["QTD_VARANDA"]]),
                self._cell_text(row[self.IDX["LBL_VARANDA"]]),
                "varanda"
            ),
            area_servico=lbl.pair_qty_label(
                self._cell_text(row[self.IDX["QTD_ASERV"]]),
                self._cell_text(row[self.IDX["LBL_ASERV"]]),
                "área de serv."
            ),
            banheiros=lbl.pair_qty_label(
                self._cell_text(row[self.IDX["QTD_BANHEIRO"]]),
                self._cell_text(row[self.IDX["LBL_BANHEIRO"]]),
                "wcs"
            ),
            salas=lbl.pair_qty_label(
                self._cell_text(row[self.IDX["QTD_SALA"]]),
                self._cell_text(row[self.IDX["LBL_SALA"]]),
                "sala"
            ),
            cozinhas=lbl.pair_qty_label(
                self._cell_text(row[self.IDX["QTD_COZINHA"]]),
                self._cell_text(row[self.IDX["LBL_COZINHA"]]),
                "cozinha"
            ),
            descricao=self._cell_text(row[self.IDX["DESCRICAO"]]),
            iptu=fmt.iptu_as_text(row[self.IDX["IPTU_VAL"]]),
            matricula=self._cell_text(row[self.IDX["MATRICULA"]]),
            oficio=self._cell_text(row[self.IDX["OFICIO"]]),
        )
