import os
import sys
import math
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side

DEFAULT_INPUT = "Teste tit2.xlsx"
DEFAULT_SHEET = "Planilha1"
OUTPUT_XLSX = "dados_formatados.xlsx"

# Layout: cards empilhados em uma única coluna (A)
COL_LETTER = "A"
COL_WIDTH_CHARS = 28
BASE_ROW_HEIGHT = 18
EXTRA_PADDING_LINES = 2
ROW_GAP = 0

# Índices (0-based) de colunas conforme a sua planilha
IDX = {
    "EMPREEND": 0,
    "TIPO": 1,
    "AREA_TOTAL": 3,
    "AREA_PRIV": 6,
    "AREA_TERR": 9,
    "QTD_QUARTO": 11,
    "LBL_QUARTO": 12,
    "VAGA": 13,
    "QTD_VARANDA": 14,
    "LBL_VARANDA": 15,
    "QTD_ASERV": 16,
    "LBL_ASERV": 17,
    "QTD_BANHEIRO": 18,
    "LBL_BANHEIRO": 19,
    "QTD_SALA": 20,
    "LBL_SALA": 21,
    "QTD_COZINHA": 22,
    "LBL_COZINHA": 23,
    "DESCRICAO": 24,   
    "IPTU_VAL": 26,    
    "MATRICULA": 28,
    "OFICIO": 30,
}

def cell_text(c):
    v = c.value if hasattr(c, "value") else c
    return "" if v is None else str(v).strip()

def format_decimal_pt(val_str: str, places: int = 2) -> str:

    if not val_str:
        return ""
    try:
        d = Decimal(str(val_str).replace(",", "."))
        q = d.quantize(Decimal(10) ** -places)  
        s = format(q, "f")
        s = s.rstrip("0").rstrip(".")
        s = s.replace(".", ",")
        return s
    except Exception:
        return val_str.replace(".", ",")

def num_to_str(val_str: str) -> str:
    return format_decimal_pt(val_str, 2)

def pair_qty_label(row, q_idx, lbl_idx, default_label: str) -> str:
    q = format_decimal_pt(cell_text(row[q_idx]), 0) 
    if not q:
        return ""
    raw = cell_text(row[lbl_idx]) or default_label
    norm_map = {
        "wc": "wc", "wcs": "wcs", "wc.": "wc",
        "quartos": "quartos", "quarto": "quarto",
        "área de serv.": "área de serv.", "area de serv.": "área de serv.", "área de serviço": "área de serv.",
        "sala": "sala", "salas": "salas", "cozinha": "cozinha", "varanda": "varanda",
        "quartos qts": "quartos"
    }
    label = norm_map.get(raw.strip().lower(), raw.strip())
    if q == "1":
        if label.endswith("s") and label not in ("wcs", "áreas de serv."):
            label = label[:-1]
        if label == "wcs":
            label = "wc"
    else:
        if label == "wc":
            label = "wcs"
    return f"{q} {label}"

def iptu_as_text(cell) -> str:
    v = cell.value if hasattr(cell, "value") else cell
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, int):
        return str(v)
    try:
        d = Decimal(str(v))
        s = format(d, "f").rstrip("0").rstrip(".")
        return s if s else "0"
    except (InvalidOperation, ValueError):
        return str(v).strip()

def montar_bloco(row):
    parts = []

    emp = cell_text(row[IDX["EMPREEND"]])
    if emp:
        parts.append(f"{emp}.")
    tipo = cell_text(row[IDX["TIPO"]])
    if tipo:
        parts.append(f"{tipo},")

    a_tot = num_to_str(cell_text(row[IDX["AREA_TOTAL"]]))
    if a_tot:
        parts.append(f"{a_tot} m\u00b2 de área total,")
    a_priv = num_to_str(cell_text(row[IDX["AREA_PRIV"]]))
    if a_priv:
        parts.append(f"{a_priv} m\u00b2 de área privativa,")
    a_terr = num_to_str(cell_text(row[IDX["AREA_TERR"]]))
    if a_terr:
        parts.append(f"{a_terr} m\u00b2 de área do terreno,")

    vaga = cell_text(row[IDX["VAGA"]])
    if vaga:
        parts.append(vaga + ",")

    varanda = pair_qty_label(row, IDX["QTD_VARANDA"], IDX["LBL_VARANDA"], "varanda")
    if varanda:
        parts.append(f"{varanda},")
    quartos = pair_qty_label(row, IDX["QTD_QUARTO"], IDX["LBL_QUARTO"], "quartos")
    if quartos:
        parts.append(f"{quartos},")
    a_serv = pair_qty_label(row, IDX["QTD_ASERV"], IDX["LBL_ASERV"], "área de serv.")
    if a_serv:
        parts.append(f"{a_serv},")
    banheiros = pair_qty_label(row, IDX["QTD_BANHEIRO"], IDX["LBL_BANHEIRO"], "wcs")
    if banheiros:
        parts.append(f"{banheiros},")
    salas = pair_qty_label(row, IDX["QTD_SALA"], IDX["LBL_SALA"], "sala")
    if salas:
        parts.append(f"{salas},")
    cozinhas = pair_qty_label(row, IDX["QTD_COZINHA"], IDX["LBL_COZINHA"], "cozinha")
    if cozinhas:
        parts.append(f"{cozinhas},")

    first_line = " ".join(parts).replace(" ,", ",").replace(",,", ",").strip()
    if first_line.endswith(","):
        first_line = first_line[:-1]
    if not first_line.endswith("."):
        first_line += "."
    first_line = first_line.replace("..", ".")

    descricao = cell_text(row[IDX["DESCRICAO"]])
    desc_block = f"\n\n{descricao}" if descricao else ""

    iptu_val = iptu_as_text(row[IDX["IPTU_VAL"]])
    matricula_val = cell_text(row[IDX["MATRICULA"]])
    oficio_val = cell_text(row[IDX["OFICIO"]])

    return f"{first_line}{desc_block}\n\nIPTU: {iptu_val}\n\nMatrícula: {matricula_val} Ofício: {oficio_val}."

def estimate_row_height(text: str, width_chars: int, base_row_height: float, pad_lines: int) -> float:
    total_lines = 0
    for para in text.split("\n"):
        total_lines += max(1, math.ceil(len(para) / max(1, width_chars))) if para else 1
    total_lines += pad_lines
    return total_lines * base_row_height

def pick_worksheet(wb, preferred_name: str | None):
    names = wb.sheetnames
    if preferred_name and preferred_name in names:
        return wb[preferred_name]
    if preferred_name:
        pref = preferred_name.strip().lower()
        for n in names:
            if n.strip().lower() == pref or pref in n.strip().lower():
                return wb[n]
    return wb.active

def processar_xlsx(input_path: str, sheet_name: str):
    wb_in = load_workbook(input_path, data_only=True, read_only=True)
    ws_in = pick_worksheet(wb_in, sheet_name)

    blocks = []
    first = True
    for row in ws_in.iter_rows(values_only=False):
        if first:
            first = False
            continue
        if not row or all(cell_text(c) == "" for c in row):
            continue
        bloco = montar_bloco(row)
        if bloco:
            blocks.append(bloco)
    wb_in.close()
    return blocks

def main():
    input_file = sys.argv[1] if len(sys.argv) >= 2 else DEFAULT_INPUT
    sheet_name = sys.argv[2] if len(sys.argv) >= 3 else DEFAULT_SHEET

    if not os.path.exists(input_file):
        print(f"Erro: O arquivo de entrada '{input_file}' não foi encontrado.")
        return

    blocks = processar_xlsx(input_file, sheet_name)
    if not blocks:
        print("Nenhum registro foi formatado.")
        return

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "Registros"
    ws.column_dimensions[COL_LETTER].width = COL_WIDTH_CHARS

    thin = Side(style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    row_ptr = 1
    for bloco in blocks:
        cell = ws[f"{COL_LETTER}{row_ptr}"]
        cell.value = bloco
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        ws.row_dimensions[row_ptr].height = estimate_row_height(
            bloco, COL_WIDTH_CHARS, BASE_ROW_HEIGHT, EXTRA_PADDING_LINES
        )
        row_ptr += 1 + ROW_GAP

    wb_out.save(OUTPUT_XLSX)
    print(f"Arquivo '{OUTPUT_XLSX}' gerado com {len(blocks)} cards em A1, A2, A3, ...")

if __name__ == "__main__":
    main()
